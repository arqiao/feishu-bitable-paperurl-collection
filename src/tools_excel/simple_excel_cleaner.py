#!/usr/bin/env python3
"""
简单的 Excel 文件空行清理工具
"""

import os

def clean_excel_file(file_path):
    """
    使用系统命令尝试修复和清理 Excel 文件
    """
    print(f"正在处理文件: {file_path}")
    
    if not os.path.exists(file_path):
        print("文件不存在")
        return False
    
    # 首先尝试备份原文件
    backup_path = file_path + '.backup'
    try:
        import shutil
        shutil.copy2(file_path, backup_path)
        print(f"已创建备份: {backup_path}")
    except Exception as e:
        print(f"创建备份失败: {e}")
        return False
    
    # 尝试使用不同的方法
    print("尝试方法 1: 使用 pandas 读取并重新保存")
    try:
        import pandas as pd
        
        # 尝试读取文件
        print("正在读取 Excel 文件...")
        
        # 使用 openpyxl 引擎显式指定
        df = pd.read_excel(file_path, engine='openpyxl')
        print(f"成功读取数据，形状: {df.shape}")
        
        # 删除空行
        print("正在删除空行...")
        df_cleaned = df.dropna(how='all')  # 删除所有列都为 NaN 的行
        print(f"清理后形状: {df_cleaned.shape}")
        print(f"删除了 {len(df) - len(df_cleaned)} 个空行")
        
        # 保存文件
        print("正在保存文件...")
        df_cleaned.to_excel(file_path, index=False, engine='openpyxl')
        print("文件保存成功！")
        return True
        
    except Exception as e:
        print(f"方法 1 失败: {e}")
    
    print("尝试方法 2: 使用 openpyxl 直接处理")
    try:
        import openpyxl
        
        print("正在加载工作簿...")
        workbook = openpyxl.load_workbook(file_path)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"处理工作表: {sheet_name}")
            
            # 标记要删除的行
            rows_to_delete = []
            
            for row_idx in range(sheet.max_row, 0, -1):
                is_empty = True
                for col_idx in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and str(cell.value).strip() != '':
                        is_empty = False
                        break
                
                if is_empty:
                    rows_to_delete.append(row_idx)
            
            # 删除空行
            for row_idx in rows_to_delete:
                sheet.delete_rows(row_idx)
            
            print(f"删除了 {len(rows_to_delete)} 个空行")
        
        print("正在保存工作簿...")
        workbook.save(file_path)
        print("文件保存成功！")
        return True
        
    except Exception as e:
        print(f"方法 2 失败: {e}")
    
    print("所有方法都失败了")
    return False

if __name__ == "__main__":
    excel_file = os.path.join("d:\\workspace\\clawbots\\feishuMSG-xls", "cfg", "WeChatGongZhongHao.xlsx")
    
    success = clean_excel_file(excel_file)
    
    if success:
        print("\nExcel 文件清理完成！")
    else:
        print("\n清理失败，请检查文件格式或手动处理。")