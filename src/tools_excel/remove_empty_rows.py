#!/usr/bin/env python3
"""
删除 Excel 文件中的空行
用于清理 cfg\WeChatGongZhongHao.xlsx 文件中的空行
"""

import openpyxl
import os

def remove_empty_rows(excel_file_path):
    """
    从 Excel 文件中删除所有空行
    
    Args:
        excel_file_path (str): Excel 文件路径
    """
    print(f"正在处理文件: {excel_file_path}")
    
    # 检查文件是否存在
    if not os.path.exists(excel_file_path):
        print(f"错误: 文件不存在 - {excel_file_path}")
        return False
    
    try:
        # 加载工作簿（尝试不同的读取模式）
        print("正在加载工作簿...")
        workbook = openpyxl.load_workbook(excel_file_path, read_only=False, keep_vba=False)
        
        # 处理所有工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"处理工作表: {sheet_name}")
            
            # 获取最大行数和列数
            max_row = sheet.max_row
            max_col = sheet.max_column
            print(f"原始行数: {max_row}, 列数: {max_col}")
            
            if max_row == 0 or max_col == 0:
                print("工作表为空，跳过处理")
                continue
            
            # 从最后一行开始向前遍历，避免删除时索引变化的问题
            rows_to_delete = []
            empty_row_count = 0
            
            for row in range(max_row, 0, -1):
                is_empty = True
                
                # 检查该行所有单元格是否为空
                for col in range(1, max_col + 1):
                    try:
                        cell_value = sheet.cell(row=row, column=col).value
                        if cell_value is not None and str(cell_value).strip() != '':
                            is_empty = False
                            break
                    except Exception as cell_error:
                        print(f"读取单元格 ({row},{col}) 时出错: {cell_error}")
                        is_empty = False  # 如果读取出错，假设不是空行
                        break
                
                if is_empty:
                    rows_to_delete.append(row)
                    empty_row_count += 1
                    
                    # 每处理1000行输出一次进度
                    if empty_row_count % 1000 == 0:
                        print(f"已发现 {empty_row_count} 个空行...")
            
            print(f"总共发现空行数: {len(rows_to_delete)}")
            
            # 删除空行
            if rows_to_delete:
                print("正在删除空行...")
                for row in rows_to_delete:
                    try:
                        sheet.delete_rows(row)
                    except Exception as delete_error:
                        print(f"删除第 {row} 行时出错: {delete_error}")
                
                print(f"成功删除 {len(rows_to_delete)} 个空行")
                print(f"清理后行数: {sheet.max_row}")
            else:
                print("未发现空行，无需删除")
        
        # 保存文件（覆盖原文件）
        print("正在保存文件...")
        workbook.save(excel_file_path)
        print(f"文件已成功保存: {excel_file_path}")
        
        return True
        
    except Exception as e:
        print(f"处理文件时出错: {str(e)}")
        import traceback
        print(f"详细错误信息: {traceback.format_exc()}")
        return False

if __name__ == "__main__":
    # Excel 文件路径
    excel_file = r"d:\workspace\clawbots\feishuMSG-xls\cfg\WeChatGongZhongHao.xlsx"
    
    # 执行空行删除
    success = remove_empty_rows(excel_file)
    
    if success:
        print("空行删除完成！")
    else:
        print("处理失败！")