#!/usr/bin/env python3
"""
简单的空行删除脚本 - 直接使用 openpyxl
"""

import os
import sys
import openpyxl

def simple_remove_empty_rows(file_path):
    """
    使用 openpyxl 直接删除空行
    """
    print(f"正在处理文件: {file_path}")
    
    if not os.path.exists(file_path):
        print("文件不存在")
        return False
    
    try:
        # 加载工作簿
        print("正在加载工作簿...")
        workbook = openpyxl.load_workbook(file_path)
        
        # 处理所有工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"处理工作表: {sheet_name}")
            
            # 确保工作表可见
            sheet.sheet_state = 'visible'
            
            # 获取最大行数
            max_row = sheet.max_row
            print(f"原始行数: {max_row}")
            
            if max_row == 0:
                print("工作表为空，跳过处理")
                continue
            
            # 从最后一行开始向前遍历，删除空行
            rows_to_delete = []
            
            for row in range(max_row, 0, -1):
                is_empty = True
                
                # 检查该行所有单元格是否为空
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value is not None and str(cell_value).strip() != '':
                        is_empty = False
                        break
                
                if is_empty:
                    rows_to_delete.append(row)
            
            # 删除空行
            if rows_to_delete:
                print(f"发现 {len(rows_to_delete)} 个空行，正在删除...")
                for row in rows_to_delete:
                    sheet.delete_rows(row)
                print(f"删除完成，剩余行数: {sheet.max_row}")
            else:
                print("未发现空行")
        
        # 保存文件
        print("正在保存文件...")
        workbook.save(file_path)
        print("文件保存成功！")
        return True
        
    except Exception as e:
        print(f"处理失败: {e}")
        import traceback
        print(f"详细错误: {traceback.format_exc()}")
        return False

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("用法: python simple_remove_empty_rows.py <excel文件路径>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    success = simple_remove_empty_rows(excel_file)
    
    if success:
        print("\n空行删除完成！")
    else:
        print("\n处理失败！")